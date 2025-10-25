import pandas as pd
from openpyxl import load_workbook

def opex_cabang_monthly(file):
    # Load workbook in streaming mode (no full memory load)
    wb = load_workbook(filename=file, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if not sheet_names:
        raise ValueError("No sheets found in the uploaded Excel file.")

    print(f"Detected {len(sheet_names)} sheets: {sheet_names}")

    # --- 1️⃣ Use the first sheet as the master list ---
    master_name = sheet_names[0]
    print(f"Using '{master_name}' as master sheet.")
    master_df = pd.read_excel(file, sheet_name=master_name)
    master_df.columns = master_df.columns.str.strip()

    # --- 2️⃣ Loop through the remaining sheets one by one ---
    for sheet_name in sheet_names[1:]:
        try:
            df = pd.read_excel(file, sheet_name=sheet_name)
            df.columns = df.columns.str.strip()

            # Identify shared columns
            common_cols = list(set(master_df.columns).intersection(df.columns))

            if not common_cols:
                print(f"⚠ Skipping sheet '{sheet_name}' (no matching columns).")
                continue

            # Merge progressively, like Excel VLOOKUP
            master_df = master_df.merge(df, on=common_cols, how="left")
            print(f"✅ Merged '{sheet_name}' using {len(common_cols)} common columns.")

        except Exception as e:
            print(f"❌ Error reading '{sheet_name}': {e}")
            continue

    wb.close()
    return master_df
